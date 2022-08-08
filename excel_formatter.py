from email.mime import image
import openpyxl
from openpyxl.styles import Alignment
import os
import argparse




IMAGE_WIDTH = 190
COLUMN_WIDTH = 25.74 #dumb
ROW_HEIGHT = 150

WORKBOOK_NAME = 'orientation_feedback.xlsx'


def Insert_Image(ws, path, anchor):
    ## anchor: cell ('A1')
    ## path: image file path'

    img = openpyxl.drawing.image.Image(path)
    img.height = IMAGE_WIDTH
    img.width= IMAGE_WIDTH
    img.anchor = anchor
    ws.add_image(img)

wb = openpyxl.load_workbook(filename = 'template.xlsx') # read an existing workbook
ws = wb.worksheets[0]

for i in range(3,100):
    ws.row_dimensions[i].height = ROW_HEIGHT / 3
ws.column_dimensions['A'].width = COLUMN_WIDTH
ws.column_dimensions['B'].width = COLUMN_WIDTH
ws.column_dimensions['C'].width = COLUMN_WIDTH
ws.column_dimensions['D'].width = COLUMN_WIDTH

image_path = 'files/Fake_connector_front.png'

image_files = sorted(os.listdir('images'))

row = 3 # starting row
for file in image_files:
    # row = 3 * (i + 2) - 4
        
    if '.png' not in file:
        continue
    if '_ISO' in file:
        column = 'A'
    elif '_SIDE' in file:
        column = 'B'
    elif '_FRONT' in file and 'oriented' not in file:
        column = 'C'
    elif '_FRONT' in file and 'oriented' in file:
        column = 'D'

    Insert_Image(ws, f"images/{file}",f"{column}{row}") # ORIENTED IMAGE
    current_cell = ws.cell(row=row, column=5)
    current_cell.value = file.replace('_FRONT','').replace('_oriented','').replace('.png','')
    current_cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells(f"A{row}:A{row + 2}")
    ws.merge_cells(f"B{row}:B{row + 2}")
    ws.merge_cells(f"C{row}:C{row + 2}")
    ws.merge_cells(f"D{row}:D{row + 2}")

    if "_oriented" in file:
        row += 3


# Insert_Image(ws, f"images/Ball Middle - V7A_oriented_FRONT.png",f"A{1}")

wb.save(WORKBOOK_NAME)

os.startfile(WORKBOOK_NAME)
